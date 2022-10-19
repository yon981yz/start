from openpyxl.styles import Font, Border, Side,  PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

#번호, 영어, 수학
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width =  5

# 1 행의 높이는 50 으로 설정
ws.row_dimensions[1].height = 30

# 스타일 적용
# 색갈 italic = 기울기 볼드 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
# 색갈 폰트 strrike 글자 중간에 줄 침 
b1.font = Font(color="CC33FF", name="Arial", strike=True)
# 색갈 글자 크기 밑줄
c1.font = Font(color="0000FF", size=20, underline="single")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90 정 넘는 셀에 대해서 초록색 적용
for row in ws.rows:
    for cell in row:
        #각 셀에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        #center, left, right, top, bottom
        #A번 번호열은 제외
        if cell.column == 1:
            continue
        # cell 이 정수형 데이터이고 90 점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")

ws.freeze_panes = "B2"


wb.save("sample_style.xlsx")