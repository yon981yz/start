# from openpyxl import load_workbook
# wb = load_workbook("sample.xlsx")
# ws = wb.active

# # 8번째 줄 비워짐
# # ws.insert_rows(8)
# # 8번째 줄 위치에 5줄 추가
# # ws.insert_rows(8, 5)

# # ws.insert_cols(2)
# # ws.insert_cols(2, 5)


# wb.save("sample_insert_rows.xlsx")

from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


# ws.delete_rows(8, 3)
ws.delete_cols(1, 2)


wb.save("sample_delete_row.xlsx")