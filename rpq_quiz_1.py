from openpyxl.styles import Font, Border, Side,  PatternFill, Alignment
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

##########################################################################################

# 표 만들기 (내식도)

ws.merge_cells("A1:I1")
ws["A1"].value = "현재까지 작성된 최종 성적 데이터"

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])

ws.append([1,10,8,5,14,26,12])
ws.append([2,7,3,7,15,24,18])
ws.append([3,9,5,8,8,12,4])
ws.append([4,7,8,7,17,21,18])
ws.append([5,7,8,7,16,25,15])
ws.append([6,3,5,8,8,17,0])
ws.append([7,4,9,10,16,27,18])
ws.append([8,6,6,6,15,19,17])
ws.append([9,10,10,9,19,30,19])
ws.append([10,9,8,8,20,25,20])

# thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
# for row in ws.rows:
#     for cell in row:
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = thin_border

for idx, cell in enumerate(ws["D"]):
        if idx == 0 or idx == 1:
                continue
        cell.value = 10
ws["H2"] = "총점"
ws["I2"] = "성적"

for idx, score in enumerate(ws["D"], start=3):
        # sum_val = sum(score[1:]) - score[3] + 10 
        ws.cell(row=idx, column=8).value = "=sum(B{}:G{})".format(idx,idx)

#         grade = None
#         if sum_val >= 90:
#                 grade = "A"
#         elif sum_val >= 80:
#                 grade = "B"
#         elif sum_val >= 70:
#                 grade = "C"
#         else:
#                 grade = "D"


#         if score[1] < 5:
#                 grade = "F"

#         ws.cell(row=idx, column=9).value = grade

# thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
# for row in ws.rows:
#     for cell in row:
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = thin_border



wb.save("sample_Quiz1.xlsx")

#########################################################################################

# 퀴즈 풀이 답안지
 
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active

# ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

# scores = [ 
# (1,10,8,5,14,26,12),
# (2,7,3,7,15,24,18),
# (3,9,5,8,8,12,4),
# (4,7,8,7,17,21,18),
# (5,7,8,7,16,25,15),
# (6,3,5,8,8,17,0),
# (7,4,9,10,16,27,18),
# (8,6,6,6,15,19,17),
# (9,10,10,9,19,30,19),
# (10,9,8,8,20,25,20)
# ]

# for s in scores:
#         ws.append(s)

# # 퀴즈2 전부 10 만들기
# for idx, cell in enumerate(ws["D"]):
#         if idx == 0:
#                 continue
#         cell.value = 10
# ws["H1"] = "총점"
# ws["I1"] = "성적"

# for idx, score in enumerate(scores, start=2):
#         sum_val = sum(score[1:]) - score[3] + 10 
#         ws.cell(row=idx, column=8).value = "=sum(B{}:G{})".format(idx,idx)

#         grade = None
#         if sum_val >= 90:
#                 grade = "A"
#         elif sum_val >= 80:
#                 grade = "B"
#         elif sum_val >= 70:
#                 grade = "C"
#         else:
#                 grade = "D"


#         if score[1] < 5:
#                 grade = "F"

#         ws.cell(row=idx, column=9).value = grade







# wb.save("sample_Quiz1_ans.xlsx")