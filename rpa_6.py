# from openpyxl import load_workbook
# from random import *
# wb = load_workbook("sample.xlsx")
# ws = wb.active

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어"

# for i in range(2,12):
#     ws.cell(row=i, column=2, value=randint(0, 100))

# # ws.move_range("C1:C11", rows=5, cols=-1)


# wb.save("sample_move.xlsx")

#차트

from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

from openpyxl.chart import BarChart, Reference, LineChart
# #B2:C11 까지의 데이터를 차트로 생성
# bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
# #차트 종류 결정 (bar, line, pie ...)
# bar_chart = BarChart()
# #차트 데이터 추가
# bar_chart.add_data(bar_value)
# #차트 넣을 위치 정의
# ws.add_chart(bar_chart, "E1")

line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3)
line_chart = LineChart()
# 계열 > 영어, 수학 (제목에서 가저옴)
line_chart.add_data(line_value, titles_from_data=True)
# 차트 제목
line_chart.title = "성적표"
# 차트 스타일
line_chart.style = 20
# 차트 각 축의 이름
line_chart.y_axis.title = "점수"
line_chart.x_axis.title = "번호"
# #차트 넣을 위치 정의
ws.add_chart(line_chart, "E1")




wb.save("sample_chart.xlsx")

