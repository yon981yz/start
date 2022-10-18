from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "test"
# wb.save("test2.xlsx")
# wb.close()

# ws = wb.create_sheet()
# ws.title = "MySheet"
# ws.sheet_properties.tabColor = "ff00ff"

# ws1 = wb.create_sheet("YourSheet")
# ws2 = wb.create_sheet("NewSheet", 2)

# new_ws = wb["NewSheet"]

# print(wb.sheetnames)

# new_ws["A1"] = "Test"
# target = wb.copy_worksheet(new_ws)
# target.title = "Copied Sheet"

# wb.save("sample.xlsx")


ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)

print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)

c = ws.cell(column=3, row=1, value=10)
print(c.value)

from random import *

index = 1
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0, 100))
        ws.cell(row=x, column=y, value=index)
        index += 1


wb.save("sample.xlsx")